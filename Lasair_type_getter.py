import openpyxl
import os, sys
import requests
import pandas as pd
from bs4 import BeautifulSoup, SoupStrainer
import traceback
url = "https://lasair-ztf.lsst.ac.uk/objects/%s/"
xlsxpath = sys.argv[1]
wb_obj = openpyxl.load_workbook(xlsxpath)
sheet_obj = wb_obj.active

session = requests.Session()
strainer = SoupStrainer('div', attrs={'class': 'col-12 col-lg-6 mb-4'})

for i in range(2, sheet_obj.max_row+1):
    try:
        cell_obj = sheet_obj.cell(row = i, column = 1)
        name = cell_obj.value
        gen_cell = sheet_obj.cell(row = i, column = 6)
        if gen_cell.value:
            #print(f"({i-1}/{sheet_obj.max_row-1}) Skipped", name)
            continue
        print(f"({i-1}/{sheet_obj.max_row-1}) Getting", name, "...")
        type_cell = sheet_obj.cell(row = i, column = 2)
        z_cell = sheet_obj.cell(row = i, column = 4)
        n_cell = sheet_obj.cell(row = i, column = 5)

        r = session.get(url%name)

        soup = BeautifulSoup(r.content, 'lxml', parse_only=strainer)
        table = soup.find('img', attrs = {'src':'/lasair/static//img/icons/tns2.png'}).parent.parent.parent.find('small')
        for row in table.findAll('b'):
            if "z =" in row.text:
                z_cell.value = float(row.text.replace("z", "").replace("=", "").replace(" ", ""))
            elif "SN" in row.text:
                type_cell.value = row.text

        data = pd.read_html(url%name)
        data = data[1]
        data = data[data['unforced mag status'] != 'limit']
        n_cell.value = len(data.index)

        gen_cell.value = 1
        wb_obj.save(sys.argv[1])
    except KeyboardInterrupt:
        exit()
    except:
        traceback.print_exc()
wb_obj.save(sys.argv[1])
print("done")
