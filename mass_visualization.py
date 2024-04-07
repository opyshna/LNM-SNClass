import pandas as pd
import numpy as np
import openpyxl
import os, io, sys
import math
import matplotlib as mpl
import matplotlib.pyplot as plt
import seaborn
import traceback

dir = os.getcwd()
if dir[-1] != "/": dir+='/'
l = [os.path.join(dp,f) for dp, dn, filenames in os.walk(dir) for f in filenames if ((os.path.splitext(f)[1] == '.txt') or (os.path.splitext(f)[1] == '.csv')) and (f[:3]=="ZTF")]
l.sort()

xlsxpath = sys.argv[1]
filter_out_lt32 = 0
wb_obj = openpyxl.load_workbook(xlsxpath)
sheet_obj = wb_obj.active
lasair_types = {}
n_points_lasair = {}
names_xlsx = [""]*(sheet_obj.max_row+1)
for i in range(2, sheet_obj.max_row+1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    name = cell_obj.value
    names_xlsx[i]=name
    type_cell = sheet_obj.cell(row = i, column = 2)
    z_cell = sheet_obj.cell(row = i, column = 4)
    n_cell = sheet_obj.cell(row = i, column = 5)
    if type_cell.value and ((not filter_out_lt32) or n_cell.value>31): lasair_types[name] = type_cell.value.strip()
    if n_cell.value: n_points_lasair[name] = int(n_cell.value)

mass_sum = {} # {name: [with_z, without_z]}; with/without_z = [summary, summaryS, summaryS3]

lasair_to_s3 = {"SN Ia":"SN Ia  ", "SN Ib":"SN Ib/c", "SN Ic":"SN Ib/c", "SN Ib/c":"SN Ib/c", "SN Ic-BL":"SN Ib/c", "SN IIL/P":"SN II  ", "SN IIP":"SN II  ", "SN IIL":"SN II  ", "SN IIn":"SN II  ", "SN II":"SN II  ", "SN II-pec":"SN II  ", "SN IIb":"SN II  "}

cumulative_m = ['hsiao', 'hsiao-subsampled', 'nugent-hyper', 'nugent-sn1a', 'nugent-sn1bc', 'nugent-sn2l', 'nugent-sn2n', 'nugent-sn2p', 'nugent-sn91bg', 'nugent-sn91t']

for f in l:
    try:
        name = f.split("/")[-1].split("_")[0].split(".")[0]
        with_z = ("_z_" in f) or (f[-3:]=="txt")
        if name not in mass_sum: mass_sum[name] = [None, None]
        if f[-3:]=="txt":
            if f[-6:]=="_z.txt":
                print(f, "is txt without redshift, skipping")
                continue
            with open(f, "r") as f2:
                summary = f2.read()
            if not summary.strip():
                print(f, "is empty, skipping")
                continue
            summary = pd.read_csv(io.StringIO(summary.replace(" -", "  -")), sep='\s\s+', engine='python')
        else:
            summary = pd.read_csv(f)

        summary["l"] = np.exp(summary["logl"])
        summary["z"] = np.exp(summary["logz"])
        summary["aicscore"] = np.exp(-summary["AIC"])

        summary_c = summary[summary["model"].isin(cumulative_m)]
        summary_s = {}
        for i in range(len(summary)):
            d = summary.iloc[i]
            key = d["type"]
            if key not in summary_s: summary_s[key] = [0, 0, 0, 0]
            summary_s[key][0] += d["l"]
            summary_s[key][1] += d["z"]
            summary_s[key][2] += d["aicscore"]
            summary_s[key][3] += 1
        summary_s = pd.DataFrame([["Σ", i[0], i[1][0], i[1][1], i[1][2], i[1][3]] for i in summary_s.items()], columns=["model", "type", "l", "z", "aicscore", "num"])

        # print_bar_chart("Likelihood, Σ", summary_s, "l", printmodel=False)
        # print_bar_chart("Evidence, Σ", summary_s, "z", printmodel=False)
        # print_bar_chart("AIC, Σ", summary_s, "aicscore", printmodel=False)

        summary_s3 = {"SN Ia  ": ["SN Ia"], "SN Ib/c": ["SN Ib", "SN Ic", "SN Ib/c", "SN Ic-BL"], "SN II  ": ["SN IIL/P", "SN IIP", "SN IIL", "SN IIn", "SN II", "SN II-pec"]}
        summary_s3 = {i[0]: [sum(summary_s[summary_s["type"].isin(i[1])]["l"]), sum(summary_s[summary_s["type"].isin(i[1])]["z"]), sum(summary_s[summary_s["type"].isin(i[1])]["aicscore"]), sum(summary_s[summary_s["type"].isin(i[1])]["num"])] for i in summary_s3.items()}
        summary_s3 = pd.DataFrame([["Σ", i[0], i[1][0], i[1][1], i[1][2], i[1][3]] for i in summary_s3.items()], columns=["model", "type", "l", "z", "aicscore", "num"])
        # print_bar_chart("Likelihood, Σ3", summary_s3, "l", printmodel=False)
        # print_bar_chart("Evidence, Σ3", summary_s3, "z", printmodel=False)
        # print_bar_chart("AIC, Σ3", summary_s3, "aicscore", printmodel=False)

        ss = [summary, summary_s, summary_s3, summary_c]
        i2 = 0 if with_z else 1
        if mass_sum[name][i2] is None: mass_sum[name][i2] = ss
        else: print("WARNING: duplicte", i2, "for '"+name+"' in file '"+f+"'.")
        if name in names_xlsx: sheet_obj.cell(row = names_xlsx.index(name), column = 10 if with_z else 11).value = 1
    except Exception as e:
        ex = traceback.format_exc()
        try:
            summary = pd.read_csv(f)
            if "time" in summary:
                print(f, "is a lightcurve, skipping")
                continue
        except:
            pass
        print(ex, "\n\nFor", f, "   skipping...")
wb_obj.save(xlsxpath)
print()

mass_sum_without_holes = mass_sum.copy()
for name, zss in mass_sum.items():
    if zss[0] is None:
        print(name, "has no data with z, purging")
        mass_sum_without_holes.pop(name)
        continue
    if zss[1] is None:
        print(name, "has no data without z, purging")
        mass_sum_without_holes.pop(name)
        continue

cmap = mpl.colormaps.get_cmap('Oranges')
diagstyle = ':C7'

# title = "\nConfusion matrix for type with z and type without z for Σ3 by likelihood (with_z/without)"
# print(title)
# fig, ax = plt.subplots()
# fig.suptitle(title)
# metric="l"
# xlabel='Σ3, likelihood, with z'
# ylabel='Σ3, likelihood, without z'
# d = {"SN Ia  ": ["SN Ia"], "SN Ib/c": ["SN Ib", "SN Ic", "SN Ib/c", "SN Ic-BL"], "SN II  ": ["SN IIL/P", "SN IIP", "SN IIL", "SN IIn", "SN II", "SN II-pec"]}
# n = len(d)
# types = [k.strip() for k,v in d.items()]
# types.sort()
# mat = [[0]*n for i in range(n)]
# tot = 0
# for name, zs in mass_sum_without_holes.items():
#     xt = zs[0][2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     x = types.index(xt.strip())
#     yt = zs[1][2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     y = types.index(yt.strip())
#     if x!=y: print(name, xt, yt)
#     mat[y][x]+=1
#     tot+=1
# print(f"total = {tot}; accuracy = {sum([mat[i][i] for i in range(n)])/tot}")
# seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
# ax.set_xlabel(xlabel)
# ax.set_ylabel(ylabel)
# plt.plot([n,0], [n,0], diagstyle)
# plt.draw()

# title = "\nConfusion matrix for type with z and type without z for Σ3 by evidence (with_z/without)"
# print(title)
# fig, ax = plt.subplots()
# fig.suptitle(title)
# metric="z"
# xlabel='Σ3, evidence, with z'
# ylabel='Σ3, evidence, without z'
# d = {"SN Ia  ": ["SN Ia"], "SN Ib/c": ["SN Ib", "SN Ic", "SN Ib/c", "SN Ic-BL"], "SN II  ": ["SN IIL/P", "SN IIP", "SN IIL", "SN IIn", "SN II", "SN II-pec"]}
# n = len(d)
# types = [k.strip() for k,v in d.items()]
# types.sort()
# mat = [[0]*n for i in range(n)]
# tot = 0
# for name, zs in mass_sum_without_holes.items():
#     xt = zs[0][2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     x = types.index(xt.strip())
#     yt = zs[1][2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     y = types.index(yt.strip())
#     if x!=y: print(name, xt, yt)
#     mat[y][x]+=1
# seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
# ax.set_xlabel(xlabel)
# ax.set_ylabel(ylabel)
# plt.plot([n,0], [n,0], diagstyle)
# plt.draw()
#
# title = "\nConfusion matrix for type with z and type without z for Σ3 by aicscore (with_z/without)"
# print(title)
# fig, ax = plt.subplots()
# metric="aicscore"
# xlabel='Σ3, aicscore, with z'
# ylabel='Σ3, aicscore, without z'
# d = {"SN Ia  ": ["SN Ia"], "SN Ib/c": ["SN Ib", "SN Ic", "SN Ib/c", "SN Ic-BL"], "SN II  ": ["SN IIL/P", "SN IIP", "SN IIL", "SN IIn", "SN II", "SN II-pec"]}
# n = len(d)
# types = [k.strip() for k,v in d.items()]
# types.sort()
# mat = [[0]*n for i in range(n)]
# tot = 0
# for name, zs in mass_sum_without_holes.items():
#     xt = zs[0][2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     x = types.index(xt.strip())
#     yt = zs[1][2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     y = types.index(yt.strip())
#     if x!=y: print(name, xt, yt)
#     mat[y][x]+=1
#     tot+=1
# print(f"total = {tot}; accuracy = {sum([mat[i][i] for i in range(n)])/tot}")
# seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
# ax.set_xlabel(xlabel)
# ax.set_ylabel(ylabel)
# plt.plot([n,0], [n,0], diagstyle)
# plt.draw()

# title = "\nConfusion matrix for type with z and type without z for Σ by likelihood (with_z/without)"
# print(title)
# fig, ax = plt.subplots()
# fig.suptitle(title)
# metric="l"
# xlabel='Σ, likelihood, with z'
# ylabel='Σ, likelihood, without z'
# d = list(mass_sum_without_holes.items())[0][1][1][1]["type"].iloc
# types = [k.strip() for k in d]
# types.sort()
# n = len(types)
# #print(types)
# mat = [[0]*n for i in range(n)]
# tot = 0
# for name, zs in mass_sum_without_holes.items():
#     xt = zs[0][1].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     x = types.index(xt.strip())
#     yt = zs[1][1].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     y = types.index(yt.strip())
#     #print(name, x, xt, y, yt)
#     if x!=y: print(name, xt, yt)
#     mat[y][x]+=1
#     tot+=1
# print(f"total = {tot}; accuracy = {sum([mat[i][i] for i in range(n)])/tot}")
# seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
# ax.set_xlabel(xlabel)
# ax.set_ylabel(ylabel)
# plt.plot([n,0], [n,0], diagstyle)
# plt.draw()

title = "\nConfusion matrix for type for Lasair and Σ by likelihood with z (Lasair/Σ)"
print(title)
fig, ax = plt.subplots()
fig.suptitle(title)
metric="l"
xlabel='Lasair'
ylabel='Σ, likelihood, with z'
d = list(mass_sum_without_holes.items())[0][1][1][1]["type"].iloc
types = [k.strip() for k in d]
types.sort()
n = len(types)
#print(types)
mat = [[0]*n for i in range(n)]
tot = 0
for name, zs in mass_sum_without_holes.items():
    xt = lasair_types.get(name)
    if not xt: continue
    x = types.index(xt.strip())
    yt = zs[0][1].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
    y = types.index(yt.strip())
    #print(name, x, xt, y, yt)
    if x!=y: print(name, xt, yt)
    mat[y][x]+=1
    tot+=1
print(f"total = {tot}; accuracy = {sum([mat[i][i] for i in range(n)])/tot}")
seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
ax.set_xlabel(xlabel)
ax.set_ylabel(ylabel)
plt.plot([n,0], [n,0], diagstyle)
plt.draw()

title = "\nConfusion matrix for type for Lasair and Σ by likelihood without z (Lasair/Σ)"
print(title)
fig, ax = plt.subplots()
fig.suptitle(title)
metric="l"
xlabel='Lasair'
ylabel='Σ, likelihood, without z'
d = list(mass_sum_without_holes.items())[0][1][1][1]["type"].iloc
types = [k.strip() for k in d]
types.sort()
n = len(types)
#print(types)
mat = [[0]*n for i in range(n)]
tot = 0
for name, zs in mass_sum_without_holes.items():
    xt = lasair_types.get(name)
    if not xt: continue
    x = types.index(xt.strip())
    yt = zs[1][1].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
    y = types.index(yt.strip())
    #print(name, x, xt, y, yt)
    if x!=y: print(name, xt, yt)
    mat[y][x]+=1
    tot+=1
print(f"total = {tot}; accuracy = {sum([mat[i][i] for i in range(n)])/tot}")
seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
ax.set_xlabel(xlabel)
ax.set_ylabel(ylabel)
plt.plot([n,0], [n,0], diagstyle)
plt.draw()

# title = "\nConfusion matrix for type for Lasair and Σ by evidence with z (Lasair/Σ)"
# print(title)
# fig, ax = plt.subplots()
# fig.suptitle(title)
# metric="z"
# xlabel='Lasair'
# ylabel='Σ, evidence, with z'
# d = list(mass_sum_without_holes.items())[0][1][1][1]["type"].iloc
# types = [k.strip() for k in d]
# types.sort()
# n = len(types)
# #print(types)
# mat = [[0]*n for i in range(n)]
# tot = 0
# for name, zs in mass_sum_without_holes.items():
#     xt = lasair_types.get(name)
#     if not xt: continue
#     x = types.index(xt.strip())
#     yt = zs[0][1].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     y = types.index(yt.strip())
#     #print(name, x, xt, y, yt)
#     if x!=y: print(name, xt, yt)
#     mat[y][x]+=1
#     tot+=1
# print(f"total = {tot}; accuracy = {sum([mat[i][i] for i in range(n)])/tot}")
# seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
# ax.set_xlabel(xlabel)
# ax.set_ylabel(ylabel)
# plt.plot([n,0], [n,0], diagstyle)
# plt.draw()
#
# title = "\nConfusion matrix for type for Lasair and Σ by evidence without z (Lasair/Σ)"
# print(title)
# fig, ax = plt.subplots()
# fig.suptitle(title)
# metric="z"
# xlabel='Lasair'
# ylabel='Σ, evidence, without z'
# d = list(mass_sum_without_holes.items())[0][1][1][1]["type"].iloc
# types = [k.strip() for k in d]
# types.sort()
# n = len(types)
# #print(types)
# mat = [[0]*n for i in range(n)]
# tot = 0
# for name, zs in mass_sum_without_holes.items():
#     xt = lasair_types.get(name)
#     if not xt: continue
#     x = types.index(xt.strip())
#     yt = zs[1][1].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
#     y = types.index(yt.strip())
#     #print(name, x, xt, y, yt)
#     if x!=y: print(name, xt, yt)
#     mat[y][x]+=1
#     tot+=1
# print(f"total = {tot}; accuracy = {sum([mat[i][i] for i in range(n)])/tot}")
# seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
# ax.set_xlabel(xlabel)
# ax.set_ylabel(ylabel)
# plt.plot([n,0], [n,0], diagstyle)
# plt.draw()

title = "\nConfusion matrix for type for Lasair and Σ3 by likelihood with z (Lasair/Σ3)"
print(title)
fig, ax = plt.subplots()
fig.suptitle(title)
metric="l"
xlabel='Lasair'
ylabel='Σ3, likelihood, with z'
d = {"SN Ia  ": ["SN Ia"], "SN Ib/c": ["SN Ib", "SN Ic", "SN Ib/c", "SN Ic-BL"], "SN II  ": ["SN IIL/P", "SN IIP", "SN IIL", "SN IIn", "SN II", "SN II-pec"]}
n = len(d)
types = [k.strip() for k,v in d.items()]
types.sort()
mat = [[0]*n for i in range(n)]
tot = 0
for name, zs in mass_sum_without_holes.items():
    xt = lasair_types.get(name)
    if not xt: continue
    xt = lasair_to_s3[xt] # failhard
    x = types.index(xt.strip())
    yt = zs[0][2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
    y = types.index(yt.strip())
    if x!=y: print(name, xt, yt)
    mat[y][x]+=1
    tot+=1
print(f"total = {tot}; accuracy = {sum([mat[i][i] for i in range(n)])/tot}")
seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
ax.set_xlabel(xlabel)
ax.set_ylabel(ylabel)
plt.plot([n,0], [n,0], diagstyle)
plt.draw()

title = "\nConfusion matrix for type for Lasair and Σ3 by likelihood without z (Lasair/Σ3)"
print(title)
fig, ax = plt.subplots()
fig.suptitle(title)
metric="l"
xlabel='Lasair'
ylabel='Σ3, likelihood, without z'
d = {"SN Ia  ": ["SN Ia"], "SN Ib/c": ["SN Ib", "SN Ic", "SN Ib/c", "SN Ic-BL"], "SN II  ": ["SN IIL/P", "SN IIP", "SN IIL", "SN IIn", "SN II", "SN II-pec"]}
n = len(d)
types = [k.strip() for k,v in d.items()]
types.sort()
mat = [[0]*n for i in range(n)]
tot = 0
for name, zs in mass_sum_without_holes.items():
    xt = lasair_types.get(name)
    if not xt: continue
    xt = lasair_to_s3[xt] # failhard
    x = types.index(xt.strip())
    yt = zs[1][2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
    y = types.index(yt.strip())
    if x!=y: print(name, xt, yt)
    mat[y][x]+=1
    tot+=1
print(f"total = {tot}; accuracy = {sum([mat[i][i] for i in range(n)])/tot}")
seaborn.heatmap(mat, cmap=cmap, annot=True, cbar=False, square=True, xticklabels=types, yticklabels=types)
ax.set_xlabel(xlabel)
ax.set_ylabel(ylabel)
plt.plot([n,0], [n,0], diagstyle)
plt.draw()



title = "\nGraph accuracy(number of points) for type for Lasair and Σ3 by likelihood without z (Lasair/Σ3) (averaging approximation)"
print(title)
fig, ax = plt.subplots()
fig.suptitle(title)
metric="l"
xlabel='# points'
ylabel='% accuracy (average)'
d = {"SN Ia  ": ["SN Ia"], "SN Ib/c": ["SN Ib", "SN Ic", "SN Ib/c", "SN Ic-BL"], "SN II  ": ["SN IIL/P", "SN IIP", "SN IIL", "SN IIn", "SN II", "SN II-pec"]}
n = len(d)
types = [k.strip() for k,v in d.items()]
types.sort()
npl = [(n, name) for name, n in n_points_lasair.items()]
npl.sort()
xx, yy = [], []
score = 0
tot = 0
totn = 0
pn=npl[0][0]
print(pn)
for n, name in npl:
    zs = mass_sum.get(name)
    if not zs: continue
    zs = zs[1]
    if not zs: continue
    xt = lasair_types.get(name)
    if not xt: continue
    xt = lasair_to_s3[xt] # failhard
    x = types.index(xt.strip())
    yt = zs[2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
    y = types.index(yt.strip())
    if n!=pn:
        if tot != 0:
            xx+=[totn/tot]
            yy+=[score/tot]
        pn=n
    if x==y: score+=1
    tot+=1
    totn+=n
xx+=[totn/tot]
yy+=[score/tot]
npl.reverse()
xxr, yyr = [], []
score = 0
tot = 0
totn = 0
pn=npl[0][0]
for n, name in npl:
    zs = mass_sum.get(name)
    if not zs: continue
    zs = zs[1]
    if not zs: continue
    xt = lasair_types.get(name)
    if not xt: continue
    xt = lasair_to_s3[xt] # failhard
    x = types.index(xt.strip())
    yt = zs[2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
    y = types.index(yt.strip())
    if n!=pn:
        if tot != 0:
            xxr+=[totn/tot]
            yyr+=[score/tot]
        pn=n
    if x==y: score+=1
    tot+=1
    totn+=n
xxr+=[totn/tot]
yyr+=[score/tot]
plt.plot(xx[10:]+xxr[:10:-1], yy[10:]+yyr[:10:-1], '.-')
plt.axvline(x=xxr[-1], color="C7", linestyle=":")
ax.set_xlabel(xlabel)
ax.set_ylabel(ylabel)
plt.draw()

title = "\nGraph accuracy(number of points) for type for Lasair and Σ3 by likelihood with z (Lasair/Σ3) (averaging approximation)"
print(title)
fig, ax = plt.subplots()
fig.suptitle(title)
metric="l"
xlabel='# points'
ylabel='% accuracy (average)'
d = {"SN Ia  ": ["SN Ia"], "SN Ib/c": ["SN Ib", "SN Ic", "SN Ib/c", "SN Ic-BL"], "SN II  ": ["SN IIL/P", "SN IIP", "SN IIL", "SN IIn", "SN II", "SN II-pec"]}
n = len(d)
types = [k.strip() for k,v in d.items()]
types.sort()
npl = [(n, name) for name, n in n_points_lasair.items()]
npl.sort()
xx, yy = [], []
score = 0
tot = 0
totn = 0
pn=npl[0][0]
print(pn)
for n, name in npl:
    zs = mass_sum.get(name)
    if not zs: continue
    zs = zs[0]
    if not zs: continue
    xt = lasair_types.get(name)
    if not xt: continue
    xt = lasair_to_s3[xt] # failhard
    x = types.index(xt.strip())
    yt = zs[2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
    y = types.index(yt.strip())
    if n!=pn:
        if tot != 0:
            xx+=[totn/tot]
            yy+=[score/tot]
        pn=n
    if x==y: score+=1
    if x!=y: print(n,types[x],types[y],name)
    tot+=1
    totn+=n
xx+=[totn/tot]
yy+=[score/tot]
npl.reverse()
xxr, yyr = [], []
score = 0
tot = 0
totn = 0
pn=npl[0][0]
for n, name in npl:
    zs = mass_sum.get(name)
    if not zs: continue
    zs = zs[0]
    if not zs: continue
    xt = lasair_types.get(name)
    if not xt: continue
    xt = lasair_to_s3[xt] # failhard
    x = types.index(xt.strip())
    yt = zs[2].sort_values(by=[metric], ascending=False, inplace=False)["type"].iloc[0]
    y = types.index(yt.strip())
    if n!=pn:
        if tot != 0:
            xxr+=[totn/tot]
            yyr+=[score/tot]
        pn=n
    if x==y: score+=1
    tot+=1
    totn+=n
xxr+=[totn/tot]
yyr+=[score/tot]
plt.plot(xx[10:]+xxr[:10:-1], yy[10:]+yyr[:10:-1], '.-')
plt.axvline(x=xxr[-1], color="C7", linestyle=":")
ax.set_xlabel(xlabel)
ax.set_ylabel(ylabel)
plt.draw()

plt.show()
