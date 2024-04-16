#%%file pool.py
# then 
# import sys
# !{sys.executable} pool.py
import pandas as pd
import openpyxl
import sncosmo
from redback.get_data import get_lasair_data
import logging
logging.getLogger('redback').setLevel(40) # just shut up
import traceback
from progress.bar import ChargingBar
from threading import Thread
import time
import multiprocessing
import signal
import math
import functools
import numpy as np
import sys, os, signal
# variables
num_threads = multiprocessing.cpu_count() # take just enough for 100% CPU usage. More means faster if below 100% and always more RAM usage
transients = []

xlsxpath = sys.argv[1]
wb_obj = openpyxl.load_workbook(xlsxpath)
sheet_obj = wb_obj.active
max_row = 1
for i in range(2, sheet_obj.max_row+1):
    if sheet_obj.cell(row = i, column = 1).value is None:
        max_row = i-1
        break
do_cells = [None]*len(transients)
done_cells = [None]*len(transients)
for i in range(2, max_row+1):
    name = sheet_obj.cell(row = i, column = 1).value
    z = sheet_obj.cell(row = i, column = 4).value
    do_with_z_cell = sheet_obj.cell(row = i, column = 8)
    do_without_z_cell = sheet_obj.cell(row = i, column = 9)
    if do_without_z_cell.value:
        do_cells+=[do_without_z_cell]
        done_cells+=[sheet_obj.cell(row = i, column = 11)]
        transients+=[name]
    if do_with_z_cell.value and z:
        do_cells+=[do_with_z_cell, None]
        done_cells+=[sheet_obj.cell(row = i, column = 10), None]
        transients+=[name, z]


def test_worker(i, event):
    if event:
        event.get()
    return os.getpid(), i
def get_ids(pool, num_threads, m):
    thread_dict = {}
    # Get thread ids
    events = [m.Queue() for _ in range(num_threads - 1)]
    # These first num_threads - 1 tasks will wait until we set their events
    results = [pool.apply_async(test_worker, args=(i, event)) for i, event in enumerate(events)]
    # This last one is not passed an event and so it does not wait.
    # When it completes, we can be sure the other tasks, which have been submitted before it
    # have already been picked up by the other threads in the pool.
    id, index = pool.apply(test_worker, args=(num_threads - 1, None))
    thread_dict[str(id)] = index+1
    # let the others complete:
    for event in events:
        event.put(1)
    for result in results:
        id, index = result.get()
        thread_dict[str(id)] = index+1
    assert(len(thread_dict) == num_threads)
    return thread_dict

models = ['salt2', 'hsiao', 'hsiao-subsampled', 'nugent-hyper', 'nugent-sn1a', 'nugent-sn1bc', 'nugent-sn2l', 'nugent-sn2n', 'nugent-sn2p', 'nugent-sn91bg', 'nugent-sn91t', 's11-2004hx', 's11-2005gi', 's11-2005hl', 's11-2005hm', 's11-2005lc', 's11-2006fo', 's11-2006jl', 's11-2006jo', 'snana-04d1la', 'snana-04d4jv', 'snana-2004fe', 'snana-2004gq', 'snana-2004gv', 'snana-2004hx', 'snana-2004ib', 'snana-2005gi', 'snana-2005hm', 'snana-2006ep', 'snana-2006ez', 'snana-2006fo', 'snana-2006gq', 'snana-2006iw', 'snana-2006ix', 'snana-2006jl', 'snana-2006jo', 'snana-2006kn', 'snana-2006kv', 'snana-2006lc', 'snana-2006ns', 'snana-2007iz', 'snana-2007kw', 'snana-2007ky', 'snana-2007lb', 'snana-2007ld', 'snana-2007lj', 'snana-2007ll', 'snana-2007lx', 'snana-2007lz', 'snana-2007md', 'snana-2007ms', 'snana-2007nc', 'snana-2007nr', 'snana-2007nv', 'snana-2007og', 'snana-2007pg', 'snana-2007y', 'snana-sdss004012', 'snana-sdss014475', 'snf-2011fe', 'v19-1987a', 'v19-1987a-corr', 'v19-1993j', 'v19-1993j-corr', 'v19-1994i', 'v19-1994i-corr', 'v19-1998bw', 'v19-1998bw-corr', 'v19-1999dn', 'v19-1999dn-corr', 'v19-1999em', 'v19-1999em-corr', 'v19-2002ap', 'v19-2002ap-corr', 'v19-2004aw', 'v19-2004aw-corr', 'v19-2004et', 'v19-2004et-corr', 'v19-2004fe', 'v19-2004fe-corr', 'v19-2004gq', 'v19-2004gq-corr', 'v19-2004gt', 'v19-2004gt-corr', 'v19-2004gv', 'v19-2004gv-corr', 'v19-2005bf', 'v19-2005bf-corr', 'v19-2005hg', 'v19-2005hg-corr', 'v19-2006aa', 'v19-2006aa-corr', 'v19-2006ep', 'v19-2007gr', 'v19-2007gr-corr', 'v19-2007od', 'v19-2007od-corr', 'v19-2007pk', 'v19-2007pk-corr', 'v19-2007ru', 'v19-2007ru-corr', 'v19-2007uy', 'v19-2007uy-corr', 'v19-2007y', 'v19-2007y-corr', 'v19-2008aq', 'v19-2008aq-corr', 'v19-2008ax', 'v19-2008ax-corr', 'v19-2008bj', 'v19-2008bj-corr', 'v19-2008bo', 'v19-2008bo-corr', 'v19-2008d', 'v19-2008d-corr', 'v19-2008fq', 'v19-2008fq-corr', 'v19-2008in', 'v19-2008in-corr', 'v19-2009bb', 'v19-2009bb-corr', 'v19-2009bw', 'v19-2009bw-corr', 'v19-2009dd', 'v19-2009dd-corr', 'v19-2009ib', 'v19-2009ib-corr', 'v19-2009ip', 'v19-2009ip-corr', 'v19-2009iz', 'v19-2009iz-corr', 'v19-2009jf', 'v19-2009jf-corr', 'v19-2009kr', 'v19-2009kr-corr', 'v19-2009n', 'v19-2009n-corr', 'v19-2010al', 'v19-2010al-corr', 'v19-2011bm', 'v19-2011bm-corr', 'v19-2011dh', 'v19-2011dh-corr', 'v19-2011ei', 'v19-2011ei-corr', 'v19-2011fu', 'v19-2011fu-corr', 'v19-2011hs', 'v19-2011hs-corr', 'v19-2011ht', 'v19-2011ht-corr', 'v19-2012a', 'v19-2012a-corr', 'v19-2012ap', 'v19-2012ap-corr', 'v19-2012au', 'v19-2012au-corr', 'v19-2012aw', 'v19-2012aw-corr', 'v19-2013ab', 'v19-2013ab-corr', 'v19-2013am', 'v19-2013am-corr', 'v19-2013by', 'v19-2013by-corr', 'v19-2013df', 'v19-2013df-corr', 'v19-2013ej', 'v19-2013ej-corr', 'v19-2013fs', 'v19-2013fs-corr', 'v19-2013ge', 'v19-2013ge-corr', 'v19-2014g', 'v19-2014g-corr', 'v19-2016bkv', 'v19-2016bkv-corr', 'v19-2016gkg', 'v19-2016gkg-corr', 'v19-2016x', 'v19-2016x-corr', 'v19-asassn14jb', 'v19-asassn14jb-corr', 'v19-asassn15oz', 'v19-asassn15oz-corr', 'v19-iptf13bvn', 'v19-iptf13bvn-corr', 'whalen-z15b', 'whalen-z15d', 'whalen-z15g', 'whalen-z25b', 'whalen-z25d', 'whalen-z25g', 'whalen-z40b', 'whalen-z40g'\
                                            ]
num_models = len(models)
thread_prefix = f"%(c){len(str(num_models))}s Thread %(id){len(str(num_threads))}d: " if 1 else f"\33[2KThread %(id){len(str(num_threads))}d: "
t_tot = len(list(filter(lambda x: type(x) is str, transients)))
t_curr = 0
index = 0
while index < len(transients):
    this_index = index
    guess_red_shift = 0
    red_shift = 0.065
    transient = transients[index]
    thread_dict = {}
    bar_suffix = f'%(index){len(str(num_models))}d/{len(models)} [ {(t_curr:=t_curr+1)} / {t_tot} ] %(elapsed_td)s  ETA: %(eta_td)s'
    def consumer(in_q, kill, pool):
        bar = ChargingBar(transient, max = len(models), suffix = bar_suffix)
        print('\n'+'\n'.join(["\33[2K"+(thread_prefix + "starting") % {'c':'', 'id':i+1} for i in range(num_threads)])+f'\033[{num_threads}A', end='')
        bar.start()
        def progress():
            c = in_q.get()
            if c == 123: bar.next()
            if c == -1: # There was an error => go to the next SN
                pool.terminate()
                kill[0]=100
                os.kill(os.getpid(), signal.SIGINT)

        while not kill[0]:
            if not in_q.empty():
                progress()
            else:
                time.sleep(0.05)
        while not in_q.empty():
                progress()
        bar.finish()


    if index+1 < len(transients) and type(transients[index+1]) is not str:
        red_shift = transients[index+1]
        index+=2
    else:
        guess_red_shift = 1
        index+=1
    kill = [0]
    try:
        data = get_lasair_data(transient=transient, transient_type='supernova')
    
        sncosmo_data = pd.DataFrame()
        sncosmo_data["time"] = data["time"]
        sncosmo_data["band"] = data["band"]
        sncosmo_data["flux"] = data["flux_density(mjy)"]
        sncosmo_data["flux_err"] = data["flux_density_error"]
        sncosmo_data["zp"] = [25] * len(data)
        sncosmo_data["zpsys"] = data["system"].str.lower()
        data = dict(map(lambda col_kv: (col_kv[0], list(map(lambda row_kv: row_kv[1], col_kv[1].items()))), sncosmo_data.items()))
    
        summary = ["model", "type", "amplitude", "t0", "logz", "logl", "AIC", "score(logz)", "score(logl)"]
        summary = dict([(i, []) for i in summary])
    
        x0x1c = ['salt2'] # for lookup purposes

        def analyze_model(msource_packed, shared_q, thread_dict):
          try:
            task_count, msource = msource_packed
            task_count+=1
            thread_id = thread_dict.get(str(os.getpid()))
            if not thread_id:
                time.sleep(1)
                print("!!!!", thread_dict, str(os.getpid()))
            summary1 = {}
            summary1["model"] = msource
            model = sncosmo.Model(source=msource)
            model.set(z=red_shift)
            type = "N/A"
            for m in sncosmo.models._SOURCES.get_loaders_metadata():
                if msource == m["name"]:
                    type = m["type"]
            summary1["type"] = type
    
            salt = msource in x0x1c
    
            # run the fit
            bounds={'z':(0.0001, 0.2)} if guess_red_shift else {}
            fparams = ['z'] if guess_red_shift else []
            fparams+= ['t0', 'amplitude'] if not salt else ['t0', 'x0', 'x1', 'c']
            if salt:
                print((f'\n\033[{thread_id-1}B' if thread_id>1 else '\n')+'\33[2K'+((thread_prefix + f"fitting {msource}") % {'c':task_count, 'id':thread_id}) + f'\033[{thread_id}A', end='')
                result, fitted_model = sncosmo.fit_lc(
                    data, model,
                    fparams, bounds=bounds)
                p = result.parameters
                bounds |= {'x0':(0, p[2]*10), 'x1':(p[3],p[3]), 'c': (p[4],p[4])}
            print((f'\n\033[{thread_id-1}B' if thread_id>1 else '\n')+'\33[2K'+((thread_prefix + f"nesting {msource}") % {'c':task_count, 'id':thread_id}) + f'\033[{thread_id}A', end='')
            result, fitted_model = sncosmo.nest_lc(
                data, model,
                fparams, bounds=bounds, guess_amplitude_bound=not salt)
    
    
            summary1["logz"] = result.logz
            if not salt:
                summary1["amplitude"] = str(result.param_dict["amplitude"]) + " ± " + str(result.errors["amplitude"])
            else:
                summary1["amplitude"] = "x0=" + str(result.param_dict["x0"]) + " ± " + str(result.errors["x0"]) + "; x1=" + str(result.param_dict["x1"]) + " ± " + str(result.errors["x1"]) + "; c=" + str(result.param_dict["c"]) + " ± " + str(result.errors["c"])
            summary1["t0"] = str(result.param_dict["t0"]) + " ± " + str(result.errors["t0"])
    
    
            print((f'\n\033[{thread_id-1}B' if thread_id>1 else '\n')+'\33[2K'+((thread_prefix + f" 0-area {msource}") % {'c':task_count, 'id':thread_id}) + f'\033[{thread_id}A', end='')
            # create a model ================== 2
            bounds2={}
            sys.modules["Negfix"] = 1
            for p in fparams:
                bounds2[p]=(result.param_dict[p],result.param_dict[p])
            model.set(z=red_shift)
            # run the fit
            result, fitted_model = sncosmo.nest_lc(
                data, model,
                fparams,
                bounds=bounds2)
            sys.modules["Negfix"] = 0
    
            summary1["logl"] = result.logl[0]
            summary1["AIC"] = 2*result.ndof-2*result.logl[0]
    
            shared_q.put(123)
            task_count+=1
            print((f'\n\033[{thread_id-1}B' if thread_id>1 else '\n')+'\33[2K'+((thread_prefix + "waiting") % {'c':"", 'id':thread_id}) + f'\033[{thread_id}A', end='')
            return summary1
          except:
            print((f'\n\033[{thread_id-1}B' if thread_id>1 else '\n')+''+((thread_prefix + f"!") % {'c':task_count, 'id':thread_id}) + f'\033[{thread_id}A', end='')
            print(f'\033[{num_threads}B')
            traceback.print_exc()
            print("MODEL", msource, "for", transient, "caused an exception !!!!!!!!!!!!!!!")
            shared_q.put(-1)
            return None

        m = multiprocessing.Manager()
        shared_q = m.Queue()
        original_sigint_handler = signal.signal(signal.SIGINT, signal.SIG_IGN)
        pool = multiprocessing.Pool(num_threads)
        signal.signal(signal.SIGINT, original_sigint_handler)
        thread_dict = get_ids(pool, num_threads, m)
        t1 = Thread(target = consumer, args =(shared_q, kill, pool))
        t1.start()
        summary0 = list(map(lambda x: x.get(timeout=None), [pool.apply_async(analyze_model,  args) for args in zip(enumerate(models), [shared_q]*len(models), [thread_dict]*len(models))]))
        kill[0]=1
        # collect results and fill in the blanks
        summary = {}
        for i, d in enumerate(summary0):
            if d is None: time.sleep(0.5)
            for k in d.keys():
                if not k in summary:
                    summary[k]=[type(d[k])()] * i
                if len(summary[k])<i:
                    summary[k]+=[type(d[k])()] * (i-len(summary[k]))
                if len(summary[k])>i:
                    raise Exception(f'Somehow got more rows than had data ({len(summary[k])} > {i})')
                summary[k]+=[d[k]]
            for k in summary.keys():
                if len(summary[k])<(i+1):
                    summary[k]+=[type(summary[k][0])()] * (i-len(summary[k]))
                if len(summary[k])>(i+1):
                    raise Exception(f'Somehow got more rows than had data ({len(summary[k])} > {i})')


        ltot = functools.reduce(lambda a, b: np.logaddexp(a,b), summary["logl"])
        summary["score(logl)"] = [math.exp(i-ltot) for i in summary["logl"]]
        ztot = functools.reduce(lambda a, b: np.logaddexp(a,b), summary["logz"])
        summary["score(logz)"] = [math.exp(i-ztot) for i in summary["logz"]]

        summary = pd.DataFrame(summary).sort_values(by=['score(logl)'], ascending=False)
        #print(summary.to_string())
        summary.to_csv(transient+("_z" if not guess_red_shift else "")+"_guessed_"+summary["type"].iloc[0].replace("/", "")+".csv", index=False)
        summary.to_excel(transient+("_z" if not guess_red_shift else "")+"_guessed_"+summary["type"].iloc[0].replace("/", "")+".xlsx")
        if do_cells[this_index] is not None: do_cells[this_index].value = ""
        if done_cells[this_index] is not None: done_cells[this_index].value = 1
        wb_obj.save(xlsxpath)
    except KeyboardInterrupt:
        print(f'\033[{num_threads}B')
        if kill[0]==100:
            continue
        kill[0]=1
        time.sleep(0.1)
        exit()
    except:
        print(f'\033[{num_threads}B')
        kill[0]=1
        traceback.print_exc()
        print(f"!!!!!!!!!!!! {transient} caused an exception!!!!!!!!!!!!!")
wb_obj.save(xlsxpath)
print("DONE")
